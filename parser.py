import asyncio
import urllib.parse
import openpyxl
import os
import re
import sys
import random
import time
import json
from playwright.async_api import async_playwright, Error as PlaywrightError
import warnings

warnings.filterwarnings("ignore", message="coroutine 'process_firm' was never awaited")

# --- ИСПРАВЛЕНИЕ ОШИБКИ ПРИ НАЖАТИИ CTRL+C НА WINDOWS ---
if sys.platform.startswith('win'):
    from asyncio.proactor_events import _ProactorBasePipeTransport

    def silence_event_loop_closed(func):
        def wrapper(self, *args, **kwargs):
            try:
                return func(self, *args, **kwargs)
            except (RuntimeError, ValueError) as e:
                if str(e) not in ['Event loop is closed', 'I/O operation on closed pipe']:
                    raise
        return wrapper

    _ProactorBasePipeTransport.__del__ = silence_event_loop_closed(_ProactorBasePipeTransport.__del__)
# ---------------------------------------------------------

SAVE_EVERY_N = 5
MAX_FIRM_RETRIES = 2          # сколько раз пере-пытаться карточку, если она упала с ошибкой
MAX_PAGE_RETRIES = 3          # сколько раз пере-пытаться загрузку страницы выдачи


def get_firm_id(url):
    if not url:
        return None
    match = re.search(r'/firm/(\d+)', str(url))
    return match.group(1) if match else None


def get_progress_bar(iteration, total, length=20):
    if total == 0:
        return f"[{'░' * length}] 0.0%"
    percent = ("{0:.1f}").format(100 * (iteration / float(total)))
    filled_length = int(length * iteration // total)
    bar = '█' * filled_length + '░' * (length - filled_length)
    return f"[{bar}] {percent}%"


def generate_grid(lat_min, lat_max, lon_min, lon_max, steps):
    lat_step = (lat_max - lat_min) / steps
    lon_step = (lon_max - lon_min) / steps
    points = []

    if steps == 5:
        grid_districts = [
            ["Внуково, Московский", "Теплый Стан, Коньково", "Бутово, Ясенево, Чертаново Юж.", "Бирюлево, Царицыно", "Зябликово, Братеево, Капотня"],
            ["Очаково, Солнцево", "Пр-т Вернадского, Раменки", "Чертаново Сев., Зюзино", "Нагатино, Печатники, Текстильщики", "Марьино, Люблино, Кузьминки"],
            ["Кунцево, Крылатское", "Филевский парк, Хамовники", "ЦАО (Арбат, Якиманка, Замоскворечье)", "Таганский, Басманный, Лефортово", "Перово, Новогиреево, Рязанский"],
            ["Строгино, Митино", "Хорошево-Мневники, Сокол", "Тверской, Пресня, Марьина Роща", "Сокольники, Алексеевский", "Измайлово, Гольяново"],
            ["Куркино, Сев. Тушино", "Ховрино, Головинский", "Отрадное, Коптево, Тимирязевский", "ВДНХ, Останкино, Свиблово", "Медведково, Бабушкинский"]
        ]
    else:
        grid_districts = None

    for i in range(steps):
        for j in range(steps):
            lat = round(lat_min + lat_step * (i + 0.5), 6)
            lon = round(lon_min + lon_step * (j + 0.5), 6)

            col_letter = chr(65 + j)
            row_num = steps - i

            if grid_districts:
                districts = grid_districts[i][j]
                sector_name = f"Сектор {col_letter}{row_num} ({districts})"
            else:
                sector_name = f"Сектор {col_letter}{row_num}"

            points.append((lat, lon, sector_name))

    points.sort(key=lambda x: x[2])
    return points


async def bypass_museum(page):
    if "museum" in page.url:
        try:
            btn = page.get_by_text("Пропустить обновление")
            if await btn.count() > 0:
                await btn.first.click()
                await page.wait_for_timeout(random.randint(1000, 2000))
        except Exception:
            pass


def normalize_text(t):
    return re.sub(r'\s+', ' ', (t or '')).strip().lower()


async def get_rubric_text(page):
    """
    Пытается вытащить текст рубрики/категории карточки заведения.
    2ГИС обычно кладёт рубрику рядом с h1 (часто это первый <a> или <span>
    сразу после заголовка, либо блок с классом, содержащим 'Rubric'/'rubric').
    Если ничего специфичного не нашлось — берём текст верхней части карточки
    (первые ~400 символов) как fallback для поиска ключевых слов.
    """
    try:
        rubric = await page.evaluate("""() => {
            // 1) попытка найти явный блок рубрики по атрибутам/классам
            const candidates = Array.from(document.querySelectorAll('[class*="rubric" i], [class*="Rubric" i], a[href*="/rubric/"]'));
            for (const el of candidates) {
                const txt = (el.innerText || '').trim();
                if (txt && txt.length < 120) return txt;
            }
            // 2) fallback: берём текст первого блока карточки целиком
            const main = document.querySelector('main') || document.body;
            return (main.innerText || '').slice(0, 600);
        }""")
        return normalize_text(rubric)
    except Exception:
        return ""


def matches_keywords(rubric_text, title_text, keywords):
    """
    Проверяет, относится ли карточка к нужной рубрике: ищем совпадение
    любого ключевого слова из keywords в тексте рубрики ИЛИ в названии.
    Если keywords пуст — фильтрация отключена (пропускаем всё).
    """
    if not keywords:
        return True
    haystack = f"{rubric_text} {title_text}".lower()
    return any(kw.lower().strip() in haystack for kw in keywords if kw.strip())


async def process_firm(context, firm_id, url, ws, wb, file_path, lock, semaphore, state, keywords):
    """
    Параллельная обработка карточки заведения.
    Возвращает True при успешном сохранении, False при ошибке/фильтрации —
    это нужно вызывающему коду, чтобы решить, помечать ли firm_id как
    окончательно обработанный или дать шанс на повтор.
    """
    async with semaphore:
        page = None
        for attempt in range(MAX_FIRM_RETRIES + 1):
            try:
                await asyncio.sleep(random.uniform(0.1, 0.5))

                page = await context.new_page()
                await page.route("**/*", lambda route: route.abort()
                    if route.request.resource_type in ["image", "stylesheet", "media", "font"]
                    else route.continue_())

                await page.goto(url, wait_until="domcontentloaded", timeout=25000)
                await bypass_museum(page)

                h1 = page.locator('h1')
                title = (await h1.first.inner_text()).split('\n')[0].strip() if await h1.count() > 0 else "Без названия"

                # --- ПРОВЕРКА РУБРИКИ: отсекаем карточки не по теме запроса ---
                rubric_text = await get_rubric_text(page)
                if not matches_keywords(rubric_text, title, keywords):
                    async with lock:
                        state["filtered_out"] = state.get("filtered_out", 0) + 1
                    await page.close()
                    return False
                # ----------------------------------------------------------------

                geo = page.locator('a[href*="/geo/"]')
                address = (await geo.first.inner_text()).replace('\n', ', ').strip() if await geo.count() > 0 else "Не указан"

                await page.evaluate("""() => {
                    const btns = Array.from(document.querySelectorAll('button, div, span'));
                    btns.forEach(b => {
                        const t = (b.innerText || '').toLowerCase();
                        if (t.includes('показать телефон') || t.includes('соцсети') || t.includes('показать контакт') || t.includes('ещё контакт')) {
                            try { b.click(); } catch(e) {}
                        }
                    });
                }""")
                await page.wait_for_timeout(random.randint(500, 1000))

                tels = page.locator('a[href^="tel:"]')
                phones = []
                for i in range(await tels.count()):
                    href = await tels.nth(i).get_attribute('href')
                    if href:
                        num = href.replace('tel:', '').strip()
                        if num and num not in phones:
                            phones.append(num)
                phone_str = ", ".join(phones) if phones else "Не указан"

                raw_data = await page.evaluate("""() => {
                    let results = [];
                    document.querySelectorAll('a').forEach(a => { if (a.href) results.push(a.href); });
                    document.querySelectorAll('div, span, a').forEach(el => {
                        let txt = el.innerText;
                        if (txt) {
                            txt = txt.trim();
                            if (txt.length > 4 && txt.length < 70 && !txt.includes('\\n') && !txt.includes(' ')) {
                                results.push(txt);
                            }
                        }
                    });
                    return Array.from(new Set(results));
                }""")

                websites = []
                bad_domains = ["2gis", "yandex", "google", "flamp", "apple", "play.google", "otello", "restoclub", "tomesto", "zoon", "sbermarket", "megamarket", "delivery-club", "eda.yandex", "mos.ru", "gosuslugi", "w3.org", "github.com", "yoo.money"]
                good_indicators = [".ru", ".com", ".рф", ".org", ".net", ".info", ".su", ".pro", ".moscow", ".agency", ".media", ".digital", "vk.com", "t.me", "wa.me", "taplink.cc"]

                for val in raw_data:
                    val_lower = val.lower()
                    if "redirect?url=" in val_lower:
                        parsed = urllib.parse.parse_qs(urllib.parse.urlparse(val).query)
                        if "url" in parsed:
                            val = urllib.parse.unquote(parsed["url"][0])
                            val_lower = val.lower()

                    if any(ind in val_lower for ind in good_indicators) and not any(bd in val_lower for bd in bad_domains):
                        if "@" not in val:
                            if not val_lower.startswith("http"):
                                val = "https://" + val
                            if val not in websites:
                                websites.append(val)

                real_sites = [w for w in websites if not any(s in w.lower() for s in ['vk.com', 't.me', 'instagram.com', 'wa.me', 'whatsapp.com'])]
                socials = [w for w in websites if any(s in w.lower() for s in ['vk.com', 't.me', 'instagram.com', 'wa.me', 'whatsapp.com'])]
                site_str = real_sites[0] if real_sites else (socials[0] if socials else "Нет сайта")

                site_label = "[WEB]"
                if "t.me" in site_str or "tg://" in site_str:
                    site_label = "[TG]"
                elif "vk.com" in site_str:
                    site_label = "[VK]"
                elif "wa.me" in site_str or "whatsapp" in site_str:
                    site_label = "[WA]"
                elif site_str == "Нет сайта":
                    site_label = "[---]"

                async with lock:
                    state["count"] += 1
                    curr_no = state["count"]

                    if phone_str != "Не указан":
                        state["phones_found"] = state.get("phones_found", 0) + 1
                    if site_str != "Нет сайта":
                        state["sites_found"] = state.get("sites_found", 0) + 1

                    ws.append([curr_no, title, address, phone_str, site_str, site_label, url])
                    print(f"   [{curr_no}] {title} | 📞 {phone_str} | {site_label} {site_str}")

                    state["unsaved"] = state.get("unsaved", 0) + 1
                    if state["unsaved"] >= SAVE_EVERY_N:
                        try:
                            wb.save(file_path)
                            state["unsaved"] = 0
                            print(f"   💾 [Excel] Данные сохранены (Всего в файле: {curr_no})")
                        except PermissionError:
                            print("   ⚠️ [Excel] Файл занят другой программой! Данные копятся в памяти.")

                return True

            except asyncio.CancelledError:
                raise
            except (PlaywrightError, Exception) as e:
                if attempt < MAX_FIRM_RETRIES:
                    await asyncio.sleep(1.5 * (attempt + 1))
                    continue
                async with lock:
                    state["errors"] = state.get("errors", 0) + 1
                return False
            finally:
                if page:
                    try:
                        await page.close()
                    except Exception:
                        pass
                    page = None

        return False


async def has_next_page(page) -> bool:
    """Проверяет наличие следующей страницы в пагинации 2ГИС."""
    try:
        await page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
        await page.wait_for_timeout(500)
    except Exception:
        pass

    next_link = page.locator('a[rel="next"]')
    if await next_link.count() > 0:
        return True

    next_btn = page.locator('a:has-text("Вперёд"), a:has-text("→"), a:has-text(">"), button:has-text("Вперёд"), button:has-text("→"), button:has-text(">")')
    if await next_btn.count() > 0:
        return True

    page_links = page.locator('a[href*="/page/"]')
    count = await page_links.count()
    if count > 0:
        numbers = []
        for i in range(count):
            href = await page_links.nth(i).get_attribute('href')
            if href:
                match = re.search(r'/page/(\d+)', href)
                if match:
                    numbers.append(int(match.group(1)))
        if numbers:
            max_page = max(numbers)
            current_url = page.url
            current_match = re.search(r'/page/(\d+)', current_url)
            current_page = int(current_match.group(1)) if current_match else 1
            if current_page < max_page:
                return True

    pagination = page.locator('.pagination, .pager, .pages')
    if await pagination.count() > 0:
        links = pagination.locator('a[href*="/page/"]')
        if await links.count() > 0:
            current_url = page.url
            current_match = re.search(r'/page/(\d+)', current_url)
            current_page = int(current_match.group(1)) if current_match else 1
            for i in range(await links.count()):
                href = await links.nth(i).get_attribute('href')
                if href:
                    match = re.search(r'/page/(\d+)', href)
                    if match:
                        page_num = int(match.group(1))
                        if page_num > current_page:
                            return True

    return False


async def get_result_cards(page):
    """
    Возвращает локатор карточек ИМЕННО из основного списка выдачи,
    по возможности исключая блоки 'Похожие места' / 'Рекомендуем' / рекламу,
    которые 2ГИС подмешивает внизу страницы или сбоку.

    Стратегия:
    1) Пробуем несколько известных контейнеров результатов (могут меняться
       у 2ГИС со временем — если верстка изменится, тут нужно будет
       поправить селекторы, подсмотрев актуальный DOM в devtools).
    2) Если ни один контейнер не найден — откатываемся на весь документ,
       но дополнительно исключаем ссылки, лежащие внутри блоков, чей
       ближайший заголовок-секция содержит слова "похож" / "рекоменд" /
       "интересн" (эвристика на случай изменения верстки).
    """
    container_selectors = [
        '[class*="_1rehe0w"]',      # часто встречающийся контейнер списка (может устареть)
        'div[class*="searchResult" i]',
        'div[class*="SearchResult" i]',
        'main',
    ]

    for sel in container_selectors:
        container = page.locator(sel).first
        if await container.count() > 0:
            cards = container.locator('a[href*="/firm/"]')
            if await cards.count() > 0:
                return cards

    return page.locator('a[href*="/firm/"]')


async def collect_valid_hrefs(page):
    """
    Достаёт href всех карточек из основного списка результатов и
    дополнительно отфильтровывает те, что визуально находятся под
    заголовками вида 'Похожие места', 'Рекомендуем', 'Вам может
    понравиться' и т.п. — если такая эвристика применима к текущей верстке.
    """
    cards = await get_result_cards(page)
    count = await cards.count()

    excluded_markers = ["похож", "рекоменд", "может понравит", "интересн", "вам понравится"]

    hrefs = []
    for i in range(count):
        el = cards.nth(i)
        href = await el.get_attribute('href')
        if not href:
            continue

        # Проверяем ближайший предшествующий заголовок секции (если есть)
        try:
            is_excluded = await el.evaluate("""(node, markers) => {
                let cur = node;
                for (let depth = 0; depth < 6 && cur; depth++) {
                    cur = cur.parentElement;
                    if (!cur) break;
                    const heading = cur.querySelector('h2, h3, [class*="title" i]');
                    if (heading) {
                        const t = (heading.innerText || '').toLowerCase();
                        if (markers.some(m => t.includes(m))) return true;
                    }
                }
                return false;
            }""", excluded_markers)
        except Exception:
            is_excluded = False

        if not is_excluded:
            hrefs.append(href)

    return hrefs


async def main():
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))

    json_path = os.path.join(BASE_DIR, 'spisok_poiska.json')
    if not os.path.exists(json_path):
        print(f"❌ Ошибка: Файл {json_path} не найден! Создайте его перед запуском.")
        return

    with open(json_path, 'r', encoding='utf-8') as f:
        data = json.load(f)
        raw_queries = data.get("queries", [])

    if not raw_queries:
        print("❌ Ошибка: Список запросов в JSON пуст.")
        return

    # Поддерживаем 2 формата queries:
    #  - список строк: ["кальянная", "кальян-бар"]
    #  - список объектов: [{"query": "кальянная", "keywords": ["кальян", "hookah"]}]
    # Если keywords не заданы — по умолчанию используем сам текст запроса
    # (разбитый на слова) как ключевые слова для проверки рубрики.
    search_queries = []
    for item in raw_queries:
        if isinstance(item, str):
            q = item
            kws = [w for w in re.split(r'\s+', q) if len(w) > 2]
        elif isinstance(item, dict):
            q = item.get("query", "").strip()
            kws = item.get("keywords") or [w for w in re.split(r'\s+', q) if len(w) > 2]
        else:
            continue
        if q:
            search_queries.append((q, kws))

    if not search_queries:
        print("❌ Ошибка: не удалось разобрать запросы из JSON.")
        return

    city = "moscow"
    file_path = os.path.join(BASE_DIR, "база_2gis.xlsx")
    progress_file = os.path.join(BASE_DIR, "progress_b2b.txt")
    sheet_title = "База"

    if not os.path.exists(progress_file):
        with open(progress_file, "w", encoding="utf-8") as f:
            f.write("")
        print(f"📄 Создан файл отслеживания прогресса: progress_b2b.txt")

    LAT_MIN, LAT_MAX = 55.1, 56.2
    LON_MIN, LON_MAX = 36.7, 38.4
    GRID_STEPS = 7
    ZOOM = 13

    CONCURRENCY_LIMIT = 4
    saved_ids = set()          # окончательно сохранённые фирмы (успех ИЛИ отфильтрованы по рубрике)
    lock = asyncio.Lock()
    semaphore = asyncio.Semaphore(CONCURRENCY_LIMIT)

    if os.path.exists(file_path):
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and len(row) >= 6:
                firm_id = get_firm_id(row[-1])
                if firm_id:
                    saved_ids.add(firm_id)
        results_count = ws.max_row - 1
        print(f"✅ Файл найден. Уже собрано уникальных карточек: {len(saved_ids)}")
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_title
        ws.append(["№", "Название", "Адрес", "Телефон", "Сайт", "Тип сайта", "URL"])
        wb.save(file_path)
        results_count = 0

    done_sectors = set()
    if os.path.exists(progress_file):
        with open(progress_file, "r", encoding="utf-8") as f:
            done_sectors = set(line.strip() for line in f if line.strip())
        print(f"📂 Загружено обработанных секторов: {len(done_sectors)}")

    state = {
        "count": results_count,
        "duplicates": 0,
        "errors": 0,
        "unsaved": 0,
        "phones_found": 0,
        "sites_found": 0,
        "filtered_out": 0,
    }
    start_time = time.time()
    queries_done_this_run = 0

    def format_eta(seconds):
        seconds = max(0, int(seconds))
        h, rem = divmod(seconds, 3600)
        m, s = divmod(rem, 60)
        if h:
            return f"{h}ч {m}м"
        return f"{m}м {s}с"

    async with async_playwright() as p:
        browser = await p.chromium.launch(
            headless=False,
            args=["--disable-blink-features=AutomationControlled"]
        )

        context = await browser.new_context(
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
            locale="ru-RU",
            timezone_id="Europe/Moscow",
            viewport={"width": 1280, "height": 720}
        )

        main_page = await context.new_page()
        grid_points = generate_grid(LAT_MIN, LAT_MAX, LON_MIN, LON_MAX, GRID_STEPS)

        try:
            for q_idx, (search_query, keywords) in enumerate(search_queries, 1):
                total_elapsed = time.time() - start_time
                avg_per_query = total_elapsed / queries_done_this_run if queries_done_this_run > 0 else 0
                remaining_queries = len(search_queries) - q_idx + 1
                eta = format_eta(remaining_queries * avg_per_query) if queries_done_this_run > 0 else "вычисляется..."

                print(f"\n==================================================")
                print(f"🔎 Запрос [{q_idx}/{len(search_queries)}]: '{search_query}'")
                print(f"🔑 Ключевые слова рубрики: {keywords}")
                print(f"📈 Прогресс запросов: {get_progress_bar(q_idx - 1, len(search_queries))} | Осталось: ~{eta}")
                print(f"==================================================")

                encoded_query = urllib.parse.quote(search_query)
                query_start_count = state["count"]
                query_start_duplicates = state["duplicates"]
                query_start_filtered = state.get("filtered_out", 0)
                query_start_phones = state.get("phones_found", 0)
                query_start_sites = state.get("sites_found", 0)
                query_start_time = time.time()

                for cell_idx, (lat, lon, sector_name) in enumerate(grid_points, 1):
                    progress_key = f"{search_query}|{sector_name}"
                    short_sector = sector_name.split(' (')[0]

                    if progress_key in done_sectors:
                        print(f"⏩ Сектор уже обработан: {short_sector}")
                        continue

                    print(f"\n📍 Сектор [{cell_idx}/{len(grid_points)}]: {sector_name}")

                    page_num = 1
                    duplicate_pages_count = 0
                    page_retry_count = 0

                    while True:
                        print(f"   📄 Страница {page_num}")

                        search_url = f"https://2gis.ru/{city}/search/{encoded_query}/page/{page_num}?m={lon}%2C{lat}%2F{ZOOM}"

                        try:
                            await main_page.goto(search_url, wait_until="domcontentloaded", timeout=30000)
                            await bypass_museum(main_page)
                            await main_page.wait_for_timeout(1500)

                            empty_scrolls = 0
                            for _ in range(8):
                                current_count = await main_page.locator('a[href*="/firm/"]').count()
                                if current_count >= 24:
                                    break

                                await main_page.evaluate("window.scrollTo(0, document.body.scrollHeight);")
                                await main_page.keyboard.press("PageDown")

                                try:
                                    await main_page.wait_for_function(
                                        f"document.querySelectorAll('a[href*=\"/firm/\"]').length > {current_count}",
                                        timeout=2000
                                    )
                                except Exception:
                                    pass

                                new_count = await main_page.locator('a[href*="/firm/"]').count()
                                if new_count == current_count:
                                    empty_scrolls += 1
                                    if empty_scrolls >= 2:
                                        break
                                else:
                                    empty_scrolls = 0

                            page_retry_count = 0

                        except Exception as e:
                            page_retry_count += 1
                            print(f"   ⚠️ Ошибка загрузки страницы {page_num}: {e}")
                            if page_retry_count > MAX_PAGE_RETRIES:
                                print(f"   🛑 Превышено число попыток загрузки страницы. Сектор завершён (ошибка).")
                                break
                            await asyncio.sleep(2 * page_retry_count)
                            continue

                        hrefs = await collect_valid_hrefs(main_page)
                        card_count = len(hrefs)
                        print(f"   🔍 Найдено карточек (после исключения 'похожих' блоков): {card_count}")

                        if card_count == 0:
                            print(f"   🛑 Карточек нет. Выдача сектора завершена.")
                            break

                        tasks = []
                        pending_ids = []
                        duplicates_on_page = 0
                        unparsed_on_page = 0

                        for href in hrefs:
                            firm_id = get_firm_id(href)
                            if not firm_id:
                                unparsed_on_page += 1
                                continue
                            if firm_id in saved_ids:
                                duplicates_on_page += 1
                                continue

                            clean = href.split('?')[0]
                            full_url = clean if clean.startswith('http') else f"https://2gis.ru{clean}"
                            pending_ids.append(firm_id)
                            tasks.append(process_firm(context, firm_id, full_url, ws, wb, file_path, lock, semaphore, state, keywords))

                        state["duplicates"] += duplicates_on_page

                        if duplicates_on_page > 0:
                            print(f"   ♻️ Дубликатов на странице: {duplicates_on_page}")
                        if tasks:
                            print(f"   ➕ Новых карточек к обработке: {len(tasks)}")

                        errors_before_page = state.get("errors", 0)
                        filtered_before_page = state.get("filtered_out", 0)

                        results = []
                        if tasks:
                            results = await asyncio.gather(*tasks, return_exceptions=True)

                        # firm_id считается окончательно "обработанным" (и не будет
                        # больше пытаться парситься) только если:
                        #  - карточка успешно сохранена (True), ИЛИ
                        #  - карточка была отфильтрована по рубрике (False, но без исключения)
                        # Если это была ошибка/исключение — firm_id НЕ добавляется в
                        # saved_ids, и при повторном запуске скрипта (после того как
                        # сектор всё же будет помечен пройденным) карточка окажется
                        # потеряна для этого сектора. Чтобы избежать этого, ошибки
                        # уже отработаны через ретраи внутри process_firm — сюда
                        # долетают только те, что не удалось получить после всех
                        # попыток. Логируем их отдельно.
                        lost_ids = []
                        for firm_id, result in zip(pending_ids, results):
                            if isinstance(result, Exception):
                                lost_ids.append(firm_id)
                                continue
                            # True (сохранено) или False (отфильтровано по рубрике) —
                            # в обоих случаях карточку больше не нужно пытаться парсить.
                            saved_ids.add(firm_id)

                        if lost_ids:
                            print(f"   ⚠️ Не удалось обработать {len(lost_ids)} карточек после ретраев — "
                                  f"они не помечены как готовые и могут быть подобраны повторно позже.")

                        page_errors = state.get("errors", 0) - errors_before_page
                        page_filtered = state.get("filtered_out", 0) - filtered_before_page
                        if page_errors > 0:
                            print(f"   ⚙️ Ошибок при обработке: {page_errors}")
                        if page_filtered > 0:
                            print(f"   🚫 Отфильтровано по рубрике (не соответствуют запросу): {page_filtered}")

                        if state["unsaved"] > 0:
                            try:
                                wb.save(file_path)
                                state["unsaved"] = 0
                            except PermissionError:
                                pass

                        next_exists = await has_next_page(main_page)
                        if not next_exists:
                            if card_count < 12:
                                print(f"   🛑 Следующая страница отсутствует, и карточек мало. Сектор выгружен.")
                                break
                            else:
                                print(f"   ⏳ Пагинация не найдена, но карточек много. Ждём 2 сек и проверяем ещё раз...")
                                await main_page.wait_for_timeout(2000)
                                next_exists = await has_next_page(main_page)
                                if not next_exists:
                                    print(f"   🛑 Следующая страница действительно отсутствует. Сектор завершён.")
                                    break
                                else:
                                    print(f"   ✅ Пагинация появилась! Продолжаем.")

                        if card_count > 0 and duplicates_on_page == card_count:
                            duplicate_pages_count += 1
                        else:
                            duplicate_pages_count = 0

                        if duplicate_pages_count >= 2:
                            print(f"   🛑 2 страницы подряд только с дубликатами. Дальше новых нет.")
                            break

                        page_num += 1

                    done_sectors.add(progress_key)
                    with open(progress_file, "a", encoding="utf-8") as f:
                        f.write(f"{progress_key}\n")
                    print(f"   ✅ Сектор {short_sector} завершён.")

                query_new = state["count"] - query_start_count
                query_dupes = state["duplicates"] - query_start_duplicates
                query_filtered = state.get("filtered_out", 0) - query_start_filtered
                query_elapsed = time.time() - query_start_time
                queries_done_this_run += 1

                q_phones = state.get("phones_found", 0) - query_start_phones
                q_sites = state.get("sites_found", 0) - query_start_sites
                ph_pct = (q_phones / query_new * 100) if query_new > 0 else 0
                st_pct = (q_sites / query_new * 100) if query_new > 0 else 0

                print(f"\n🏁 Итог по '{search_query}':")
                print(f"   ✅ Новых: {query_new} | ♻️ Дубликатов: {query_dupes} | 🚫 Отфильтровано по рубрике: {query_filtered} | ⏱️ Заняло: {query_elapsed:.0f}с")
                print(f"   📊 Качество данных: 📞 Телефоны {ph_pct:.1f}% | 🌐 Сайты {st_pct:.1f}%")

                if q_idx < len(search_queries):
                    pause_time = random.uniform(5.0, 10.0)
                    print(f"\n⏳ Ждем {pause_time:.1f} сек перед следующим запросом...")
                    await asyncio.sleep(pause_time)

        except asyncio.CancelledError:
            pass
        finally:
            if state.get("unsaved", 0) > 0:
                try:
                    wb.save(file_path)
                    state["unsaved"] = 0
                except PermissionError:
                    print("\n⚠️ Не удалось сохранить финальный xlsx — файл открыт в Excel!")

            total_elapsed = time.time() - start_time
            print(f"\n==================================================")
            print(f"📊 ИТОГО ЗА СЕССИЮ: собрано новых {state['count'] - results_count}, всего в файле {state['count']}")
            print(f"♻️ Дубликатов пропущено: {state['duplicates']} | 🚫 Отфильтровано по рубрике: {state.get('filtered_out', 0)} | ⚙️ Ошибок: {state.get('errors', 0)}")
            print(f"⏱️ Время работы: {format_eta(total_elapsed)}")
            print(f"==================================================")

            try:
                await browser.close()
            except Exception:
                pass


if __name__ == "__main__":
    try:
        asyncio.run(main())
    except KeyboardInterrupt:
        print("\n\n[!] Парсинг остановлен пользователем (Ctrl+C). Прогресс сохранен!")