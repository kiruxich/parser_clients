import asyncio
import urllib.parse
import openpyxl
import os
import re
import sys
import random
import time
from playwright.async_api import async_playwright, Error as PlaywrightError
import warnings

warnings.filterwarnings("ignore", message="coroutine 'process_firm' was never awaited")

# --- ИСПРАВЛЕНИЕ ОШИБКИ ПРИ НАЖАТИИ CTRL+C НА WINDOWS ---
if sys.platform.startswith('win'):
    from asyncio.proactor_events import _ProactorBasePipeTransport
    
    def silence_event_loop_closed(func):
        def wrapper(self, *args, **kwargs):
            try:
                return func(self, *args, **kwargs)
            except (RuntimeError, ValueError) as e:
                if str(e) not in ['Event loop is closed', 'I/O operation on closed pipe']:
                    raise
        return wrapper
        
    _ProactorBasePipeTransport.__del__ = silence_event_loop_closed(_ProactorBasePipeTransport.__del__)
# ---------------------------------------------------------

# Сохранять xlsx на диск не после КАЖДОЙ карточки, а раз в N карточек.
# С ростом файла openpyxl.save() каждый раз перезаписывает ВЕСЬ файл целиком,
# поэтому при тысячах строк сохранение после каждой строки заметно тормозит парсинг.
SAVE_EVERY_N = 5

def get_firm_id(url):
    if not url:
        return None
    match = re.search(r'/firm/(\d+)', str(url))
    return match.group(1) if match else None

async def bypass_museum(page):
    if "museum" in page.url:
        try:
            btn = page.get_by_text("Пропустить обновление")
            if await btn.count() > 0:
                await btn.first.click()
                await page.wait_for_timeout(random.randint(1000, 2000))
        except Exception:
            pass

def generate_grid(lat_min, lat_max, lon_min, lon_max, steps):
    """Генерирует координаты и сопоставляет их с реальными районами Москвы"""
    lat_step = (lat_max - lat_min) / steps
    lon_step = (lon_max - lon_min) / steps
    points = []
    
    grid_districts = [
        ["Внуково, Московский", "Теплый Стан, Коньково", "Бутово, Ясенево, Чертаново Юж.", "Бирюлево, Царицыно", "Зябликово, Братеево, Капотня"],
        ["Очаково, Солнцево", "Пр-т Вернадского, Раменки", "Чертаново Сев., Зюзино", "Нагатино, Печатники, Текстильщики", "Марьино, Люблино, Кузьминки"],
        ["Кунцево, Крылатское", "Филевский парк, Хамовники", "ЦАО (Арбат, Якиманка, Замоскворечье)", "Таганский, Басманный, Лефортово", "Перово, Новогиреево, Рязанский"],
        ["Строгино, Митино", "Хорошево-Мневники, Сокол", "Тверской, Пресня, Марьина Роща", "Сокольники, Алексеевский", "Измайлово, Гольяново"],
        ["Куркино, Сев. Тушино", "Ховрино, Головинский", "Отрадное, Коптево, Тимирязевский", "ВДНХ, Останкино, Свиблово", "Медведково, Бабушкинский"]
    ]

    for i in range(steps):
        for j in range(steps):
            lat = round(lat_min + lat_step * (i + 0.5), 6)
            lon = round(lon_min + lon_step * (j + 0.5), 6)
            
            col_letter = chr(65 + j)
            row_num = steps - i 
            
            if steps == 5:
                districts = grid_districts[i][j]
                sector_name = f"Сектор {col_letter}{row_num} ({districts})"
            else:
                sector_name = f"Сектор {col_letter}{row_num}"
                
            points.append((lat, lon, sector_name))
            
    points.sort(key=lambda x: x[2])
    return points

async def process_firm(context, firm_id, url, ws, wb, file_path, lock, semaphore, state):
    """Параллельная обработка карточки заведения"""
    async with semaphore:
        page = None
        try:
            # Небольшая случайная задержка перед стартом — раньше она стояла в главном
            # цикле ДО gather() и просто тормозила сканирование карточек, а не разносила
            # реальные запросы во времени. Здесь она реально "размазывает" старты 4
            # параллельных воркеров.
            await asyncio.sleep(random.uniform(0.1, 0.5))

            page = await context.new_page()
            await page.route("**/*", lambda route: route.abort() 
                if route.request.resource_type in ["image", "stylesheet", "media", "font"] 
                else route.continue_())

            await page.goto(url, wait_until="domcontentloaded", timeout=25000)
            await bypass_museum(page)

            h1 = page.locator('h1')
            title = (await h1.first.inner_text()).split('\n')[0].strip() if await h1.count() > 0 else "Без названия"

            geo = page.locator('a[href*="/geo/"]')
            address = (await geo.first.inner_text()).replace('\n', ', ').strip() if await geo.count() > 0 else "Не указан"

            await page.evaluate("""() => {
                const btns = Array.from(document.querySelectorAll('button, div, span'));
                btns.forEach(b => {
                    const t = (b.innerText || '').toLowerCase();
                    if (t.includes('показать телефон') || t.includes('соцсети') || t.includes('показать контакт') || t.includes('ещё контакт')) {
                        try { b.click(); } catch(e) {}
                    }
                });
            }""")
            await page.wait_for_timeout(random.randint(500, 1000))

            tels = page.locator('a[href^="tel:"]')
            phones = []
            for i in range(await tels.count()):
                href = await tels.nth(i).get_attribute('href')
                if href:
                    num = href.replace('tel:', '').strip()
                    if num and num not in phones:
                        phones.append(num)
            phone_str = ", ".join(phones) if phones else "Не указан"

            raw_data = await page.evaluate("""() => {
                let results = [];
                document.querySelectorAll('a').forEach(a => { if (a.href) results.push(a.href); });
                document.querySelectorAll('div, span, a').forEach(el => {
                    let txt = el.innerText;
                    if (txt) {
                        txt = txt.trim();
                        if (txt.length > 4 && txt.length < 70 && !txt.includes('\\n') && !txt.includes(' ')) {
                            results.push(txt);
                        }
                    }
                });
                return Array.from(new Set(results));
            }""")

            websites = []
            bad_domains = ["2gis", "yandex", "google", "flamp", "apple", "play.google", "otello", "restoclub", "tomesto", "zoon", "sbermarket", "megamarket", "delivery-club", "eda.yandex", "mos.ru", "gosuslugi", "w3.org", "github.com", "yoo.money"]
            good_indicators = [".ru", ".com", ".рф", ".org", ".net", ".info", ".su", ".pro", ".moscow", ".agency", ".media", ".digital", "vk.com", "t.me", "wa.me", "taplink.cc"]

            for val in raw_data:
                val_lower = val.lower()
                if "redirect?url=" in val_lower:
                    parsed = urllib.parse.parse_qs(urllib.parse.urlparse(val).query)
                    if "url" in parsed:
                        val = urllib.parse.unquote(parsed["url"][0])
                        val_lower = val.lower()

                if any(ind in val_lower for ind in good_indicators) and not any(bd in val_lower for bd in bad_domains):
                    if "@" not in val:
                        if not val_lower.startswith("http"):
                            val = "https://" + val
                        if val not in websites:
                            websites.append(val)

            real_sites = [w for w in websites if not any(s in w.lower() for s in ['vk.com', 't.me', 'instagram.com', 'wa.me', 'whatsapp.com'])]
            socials = [w for w in websites if any(s in w.lower() for s in ['vk.com', 't.me', 'instagram.com', 'wa.me', 'whatsapp.com'])]
            site_str = real_sites[0] if real_sites else (socials[0] if socials else "Нет сайта")

            async with lock:
                state["count"] += 1
                curr_no = state["count"]
                ws.append([curr_no, title, address, phone_str, site_str, url])
                print(f"[{curr_no}] {title} | {address} | {phone_str} | {site_str}")

                state["unsaved"] = state.get("unsaved", 0) + 1
                if state["unsaved"] >= SAVE_EVERY_N:
                    try:
                        wb.save(file_path)
                        state["unsaved"] = 0
                    except PermissionError:
                        # Скорее всего файл открыт в Excel — не роняем скрипт,
                        # просто попробуем сохранить в следующий раз.
                        print("   ⚠️ Не удалось сохранить xlsx — файл открыт в Excel/другой программе! Закройте файл, данные копятся в памяти.")

        except asyncio.CancelledError:
            pass
        except PlaywrightError:
            async with lock:
                state["errors"] = state.get("errors", 0) + 1
        except Exception:
            async with lock:
                state["errors"] = state.get("errors", 0) + 1
        finally:
            if page:
                try:
                    await page.close()
                except Exception:
                    pass

async def main():
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))
    
    search_queries = [
        "кальянная", "кальян бар", "лаундж бар", 
        "hookah lounge", "кальянный клуб",
        "магазин кальянов", "табак для кальяна", "кальян"
    ]
    city = "moscow"
    file_path = os.path.join(BASE_DIR, "кальянные_москвы.xlsx")
    progress_file = os.path.join(BASE_DIR, "progress_kalian.txt")
    sheet_title = "Кальянные"
    
    LAT_MIN, LAT_MAX = 55.55, 55.92
    LON_MIN, LON_MAX = 37.35, 37.85
    GRID_STEPS = 5 
    ZOOM = 14  
    
    MAX_PAGES_PER_CELL = 4 
    CONCURRENCY_LIMIT = 4

    saved_ids = set()
    lock = asyncio.Lock()
    semaphore = asyncio.Semaphore(CONCURRENCY_LIMIT)
    
    if os.path.exists(file_path):
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if len(row) >= 6 and row[5]:
                firm_id = get_firm_id(row[5])
                if firm_id:
                    saved_ids.add(firm_id)
        results_count = ws.max_row - 1
        print(f"Файл найден. Уже собрано уникальных карточек: {len(saved_ids)}")
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_title
        ws.append(["№", "Название", "Адрес", "Телефон", "Сайт", "URL"])
        wb.save(file_path)
        results_count = 0

    # Загрузка уже пройденных секторов
    done_sectors = set()
    if os.path.exists(progress_file):
        with open(progress_file, "r", encoding="utf-8") as f:
            done_sectors = set(line.strip() for line in f if line.strip())
        print(f"Загружен прогресс: пропущено уже готовых секторов: {len(done_sectors)}")

    state = {"count": results_count, "duplicates": 0, "errors": 0, "unsaved": 0}
    grid_points = generate_grid(LAT_MIN, LAT_MAX, LON_MIN, LON_MAX, GRID_STEPS)
    total_sectors = len(search_queries) * len(grid_points)
    start_time = time.time()
    sectors_done_this_run = 0

    def format_eta(seconds):
        seconds = max(0, int(seconds))
        h, rem = divmod(seconds, 3600)
        m, s = divmod(rem, 60)
        if h:
            return f"{h}ч {m}м"
        return f"{m}м {s}с"

    async with async_playwright() as p:
        browser = await p.chromium.launch(
            headless=False, 
            args=["--disable-blink-features=AutomationControlled"]
        )
        
        context = await browser.new_context(
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
            locale="ru-RU",
            timezone_id="Europe/Moscow",
            viewport={"width": 1280, "height": 720}
        )
        
        main_page = await context.new_page()

        try:
            for q_idx, search_query in enumerate(search_queries, 1):
                print(f"\n==================================================")
                print(f"🔎 Запрос [{q_idx}/{len(search_queries)}]: '{search_query}'")
                print(f"==================================================")
                encoded_query = urllib.parse.quote(search_query)
                query_start_count = state["count"]
                query_start_duplicates = state["duplicates"]

                for cell_idx, (lat, lon, sector_name) in enumerate(grid_points, 1):
                    progress_key = f"{search_query}|{sector_name}"
                    
                    short_sector_name = sector_name.split(' (')[0]

                    if progress_key in done_sectors:
                        print(f"⏩ Пропущен (уже готов): {search_query} | {short_sector_name}")
                        continue

                    print(f"\n📍 [{search_query}] [{cell_idx}/{len(grid_points)}]: {sector_name}")
                    print(f"📊 {short_sector_name}: всего страниц (максимум): {MAX_PAGES_PER_CELL}")

                    sector_start_time = time.time()
                    sector_start_count = state["count"]
                    sector_start_duplicates = state["duplicates"]

                    for page_num in range(1, MAX_PAGES_PER_CELL + 1):
                        print(f"   📄 {short_sector_name}: страница {page_num}")
                        
                        search_url = f"https://2gis.ru/{city}/search/{encoded_query}/page/{page_num}?m={lon}%2C{lat}%2F{ZOOM}"
                        
                        try:
                            await main_page.goto(search_url, wait_until="domcontentloaded", timeout=30000)
                            await bypass_museum(main_page)
                            
                            # Даем 2ГИС 2 секунды подгрузить свои тяжелые скрипты
                            await main_page.wait_for_timeout(2000) 
                            
                            # --- ЭМУЛЯЦИЯ ЧЕЛОВЕКА (Умный Скролл) ---
                            try:
                                first_card = main_page.locator('a[href*="/firm/"]').first
                                await first_card.wait_for(state="attached", timeout=10000)
                                await first_card.hover()
                            except Exception:
                                await main_page.mouse.move(200, 300)

                            prev_count = 0
                            empty_scrolls = 0  # Счетчик "холостых" прокруток
                            
                            for _ in range(6):
                                await main_page.mouse.wheel(0, 3000)
                                
                                # УМНОЕ ОЖИДАНИЕ: Ждем до 2 секунд, пока количество карточек физически не превысит старое
                                try:
                                    await main_page.wait_for_function(
                                        f"document.querySelectorAll('a[href*=\"/firm/\"]').length > {prev_count}", 
                                        timeout=2000
                                    )
                                except Exception:
                                    pass # Таймаут: новые карточки не появились
                                
                                current_count = await main_page.locator('a[href*="/firm/"]').count()
                                
                                # В 2ГИС на одной странице максимум 24 организации. Если дошли до лимита - сразу выходим!
                                if current_count >= 24:
                                    break
                                
                                # Проверка на зависание загрузки
                                if current_count == prev_count and current_count > 0:
                                    empty_scrolls += 1
                                    if empty_scrolls >= 2: # Даем 2ГИС две попытки (до 4 секунд лагов), прежде чем сдаться
                                        break 
                                else:
                                    empty_scrolls = 0 # Сбрасываем счетчик, если карточки успешно подгрузились
                                    
                                prev_count = current_count

                        except Exception:
                            continue

                        cards = main_page.locator('a[href*="/firm/"]')
                        card_count = await cards.count()

                        # --- ЗАЩИТА ОТ ТЕНЕВОГО БАНА И КАПЧИ ---
                        if card_count == 0:
                            captcha_indicators = await main_page.locator("text=/робот|капча|captcha/i").count()
                            
                            if captcha_indicators > 0:
                                print('\a')
                                print("\n[!!!] ВНИМАНИЕ: 2ГИС выдал капчу (теневой бан).")
                                print("[!!!] У вас есть 5 минут, чтобы решить её вручную в открытом окне браузера!")
                                
                                resolved = False
                                for sec in range(300):
                                    await asyncio.sleep(1)
                                    if await main_page.locator('a[href*="/firm/"]').count() > 0:
                                        print("\n[+] Капча успешно решена! Продолжаем сбор...")
                                        resolved = True
                                        card_count = await main_page.locator('a[href*="/firm/"]').count()
                                        break
                                    if sec > 0 and sec % 60 == 0:
                                        print(f"... осталось {5 - sec//60} мин ...")
                                
                                if not resolved:
                                    print("\n[-] Время вышло. Капча не решена. Скрипт остановлен для сохранения прогресса.")
                                    return 
                            else:
                                break

                        tasks = []
                        duplicates_on_page = 0
                        unparsed_on_page = 0
                        for i in range(card_count):
                            href = await cards.nth(i).get_attribute('href')
                            if href:
                                firm_id = get_firm_id(href)
                                if not firm_id:
                                    unparsed_on_page += 1
                                    continue
                                if firm_id not in saved_ids:
                                    clean = href.split('?')[0]
                                    full_url = clean if clean.startswith('http') else f"https://2gis.ru{clean}"
                                    saved_ids.add(firm_id)
                                    tasks.append(process_firm(context, firm_id, full_url, ws, wb, file_path, lock, semaphore, state))
                                else:
                                    duplicates_on_page += 1

                        state["duplicates"] += duplicates_on_page

                        # Явно показываем, ПОЧЕМУ на странице 0 новых карточек:
                        # реально пусто, всё уже собрано раньше (дубликаты), или не распарсились ссылки.
                        if duplicates_on_page > 0:
                            print(f"   ♻️ Дубликатов на странице: {duplicates_on_page} (всего за сессию: {state['duplicates']})")
                        if unparsed_on_page > 0:
                            print(f"   ❓ Не удалось извлечь ID у {unparsed_on_page} карточек (возможно, поменялась разметка сайта)")
                        if tasks:
                            print(f"   ➕ Новых карточек к обработке: {len(tasks)}")
                        elif duplicates_on_page == 0 and unparsed_on_page == 0:
                            print(f"   ⚠️ Найдено {card_count} карточек, но ни новых, ни дублей — странно, проверьте селекторы")

                        errors_before_page = state.get("errors", 0)
                        if tasks:
                            await asyncio.gather(*tasks, return_exceptions=True)
                        page_errors = state.get("errors", 0) - errors_before_page
                        if page_errors > 0:
                            print(f"   ⚙️ Ошибок при обработке карточек на этой странице: {page_errors} (всего за сессию: {state['errors']})")

                        # Гарантированный сброс на диск в конце каждой страницы —
                        # ограничивает, сколько новых строк можно потерять при аварийном завершении.
                        if state["unsaved"] > 0:
                            try:
                                wb.save(file_path)
                                state["unsaved"] = 0
                            except PermissionError:
                                print("   ⚠️ Не удалось сохранить xlsx — файл открыт в Excel/другой программе!")

                        # --- УМНАЯ ОСТАНОВКА (Smart Break) ---
                        if card_count < 12:
                            print(f"   🛑 Меньше 12 карточек. Сектор {short_sector_name} полностью собран.")
                            break

                    done_sectors.add(progress_key)
                    with open(progress_file, "a", encoding="utf-8") as f:
                        f.write(f"{progress_key}\n")

                    sector_new = state["count"] - sector_start_count
                    sector_dupes = state["duplicates"] - sector_start_duplicates
                    sector_elapsed = time.time() - sector_start_time
                    sectors_done_this_run += 1
                    total_elapsed = time.time() - start_time
                    avg_per_sector = total_elapsed / sectors_done_this_run
                    remaining_sectors = total_sectors - len(done_sectors)
                    eta = format_eta(remaining_sectors * avg_per_sector)
                    overall_pct = len(done_sectors) / total_sectors * 100
                    print(
                        f"   ✅ Сектор готов: новых {sector_new}, дублей {sector_dupes}, "
                        f"занял {sector_elapsed:.0f}с | Прогресс всего: {len(done_sectors)}/{total_sectors} "
                        f"({overall_pct:.1f}%), осталось примерно {eta}"
                    )

                query_new = state["count"] - query_start_count
                query_dupes = state["duplicates"] - query_start_duplicates
                print(f"\n🏁 Запрос '{search_query}' завершён: новых карточек {query_new}, дублей (уже были найдены другим запросом) {query_dupes}")

        except asyncio.CancelledError:
            pass
        finally:
            if state.get("unsaved", 0) > 0:
                try:
                    wb.save(file_path)
                    state["unsaved"] = 0
                except PermissionError:
                    print("⚠️ Не удалось сохранить финальный xlsx — файл открыт в Excel/другой программе! Данные могли не сохраниться.")

            total_elapsed = time.time() - start_time
            print(f"\n==================================================")
            print(f"📊 ИТОГО ЗА СЕССИЮ: собрано новых {state['count'] - results_count}, всего в файле {state['count']}")
            print(f"♻️ Дубликатов пропущено: {state['duplicates']} | ⚙️ Ошибок: {state.get('errors', 0)}")
            print(f"⏱️ Время работы: {format_eta(total_elapsed)}")
            print(f"==================================================")

            try:
                await browser.close()
            except Exception:
                pass

if __name__ == "__main__":
    try:
        asyncio.run(main())
    except KeyboardInterrupt:
        print("\n\n[!] Парсинг остановлен пользователем (Ctrl+C). Прогресс сохранен!")