import tkinter as tk
from tkinter import messagebox, scrolledtext
import openpyxl
import urllib.parse
import os
import platform
import time
import warnings

# Глушим предупреждения (если скрипт прервут)
warnings.filterwarnings("ignore", message="coroutine 'process_firm' was never awaited")

def get_domain(url):
    """Извлекает чистый домен из строки (игнорирует 'Нет сайта', 'www.', 'http://')"""
    if not url:
        return None
    url_str = str(url).strip().lower()
    
    if url_str in ["нет сайта", "не указан"]:
        return None
        
    if not url_str.startswith('http'):
        url_str = 'http://' + url_str
        
    try:
        parsed = urllib.parse.urlparse(url_str)
        domain = parsed.netloc
        if domain.startswith('www.'):
            domain = domain[4:]
        return domain
    except Exception:
        return None

def clean_franchises():
    # 1. Определяем директорию, в которой расположен файл скрипта
    script_dir = os.path.dirname(os.path.abspath(__file__))

    # 2. Ищем файлы .xlsx в этой папке (исключая временные файлы Excel, начинающиеся с ~$)
    excel_files = [
        f for f in os.listdir(script_dir) 
        if f.endswith('.xlsx') and not f.startswith('~$')
    ]

    if not excel_files:
        messagebox.showerror("Ошибка", "В папке со скриптом не найдено ни одного Excel файла (.xlsx).")
        return

    filename = excel_files[0]
    filepath = os.path.join(script_dir, filename)

    try:
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active

        domain_counts = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if len(row) >= 5:
                domain = get_domain(row[4])
                if domain:
                    domain_counts[domain] = domain_counts.get(domain, 0) + 1

        franchise_domains = {dom: count for dom, count in domain_counts.items() if count >= 2}

        if not franchise_domains:
            messagebox.showinfo("Готово", f"В файле '{filename}' нет сетей/франшиз (повторяющихся сайтов).")
            root.destroy()
            return

        rows_to_delete = []
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if len(row) >= 5:
                domain = get_domain(row[4])
                if domain in franchise_domains:
                    rows_to_delete.append(i)

        for i in reversed(rows_to_delete):
            ws.delete_rows(i)

        for i, row in enumerate(ws.iter_rows(min_row=2), start=1):
            row[0].value = i

        # --- АВТОМАТИЧЕСКОЕ ЗАКРЫТИЕ ЗАВИСШИХ ПРОЦЕССОВ ---
        saved = False
        try:
            wb.save(filepath)
            saved = True
        except PermissionError:
            answer = messagebox.askyesno(
                "Файл занят", 
                f"Файл '{filename}' сейчас открыт или удерживается зависшим процессом Excel.\n\n"
                "Принудительно закрыть Excel, чтобы сохранить очищенную базу?"
            )
            
            if answer:
                if platform.system() == 'Windows':
                    os.system("taskkill /f /im excel.exe >nul 2>&1")
                elif platform.system() == 'Darwin': 
                    os.system("killall 'Microsoft Excel' >/dev/null 2>&1")
                    os.system("killall 'Numbers' >/dev/null 2>&1")
                
                time.sleep(2)
                
                try:
                    wb.save(filepath)
                    saved = True
                except Exception as e:
                    messagebox.showerror("Ошибка", f"Не удалось сохранить даже после закрытия программ:\n{e}")
                    return
            else:
                return 

        if saved:
            sorted_franchises = sorted(franchise_domains.items(), key=lambda x: x[1], reverse=True)
            report_lines = [f"• {dom} (удалено филиалов: {count})" for dom, count in sorted_franchises]
            report_text = "\n".join(report_lines)
            show_results_in_main(len(rows_to_delete), len(franchise_domains), report_text)

    except Exception as e:
        messagebox.showerror("Ошибка", f"Произошла ошибка при обработке файла:\n{str(e)}")

def show_results_in_main(total_deleted, total_networks, details):
    """Меняет интерфейс главного окна для вывода красивого отчета"""
    
    # Скрываем стартовые элементы
    lbl_title.pack_forget()
    lbl_subtitle.pack_forget()
    btn.pack_forget()

    # Увеличиваем размер окна под отчет
    root.geometry("500x480")
    root.eval('tk::PlaceWindow . center')

    # Красивый заголовок результата
    lbl_res_title = tk.Label(
        root, 
        text="🎉 База успешно очищена!", 
        font=("Helvetica", 16, "bold"), 
        bg="#ECECEC", 
        fg="#2E7D32" # Темно-зеленый
    )
    lbl_res_title.pack(pady=(20, 5))

    lbl_summary = tk.Label(
        root, 
        text=f"Удалено строк: {total_deleted}   |   Сетевых компаний: {total_networks}", 
        font=("Helvetica", 12), 
        bg="#ECECEC", 
        fg="#333333"
    )
    lbl_summary.pack(pady=(0, 15))

    # Текстовое поле для списка (с современным шрифтом)
    txt_area = scrolledtext.ScrolledText(
        root, wrap=tk.WORD, font=("Menlo", 12), 
        bg="#FFFFFF", fg="#333333", borderwidth=0, highlightthickness=1, highlightbackground="#CCCCCC"
    )
    txt_area.pack(padx=20, pady=5, fill=tk.BOTH, expand=True)
    txt_area.insert(tk.END, details)
    txt_area.config(state=tk.DISABLED)

    # Кнопка закрытия программы
    btn_close = tk.Button(root, text="Отлично, закрыть", command=root.destroy, font=("Helvetica", 13))
    btn_close.pack(pady=15, ipadx=20)


# --- Настройка главного графического окна ---
root = tk.Tk()
root.title("Очистка базы")
root.geometry("380x220")
root.configure(bg="#ECECEC") # Цвет фона в стиле macOS
root.eval('tk::PlaceWindow . center')

# UI Элементы стартового окна
lbl_title = tk.Label(
    root, text="Очистка от франшиз", 
    font=("Helvetica", 18, "bold"), bg="#ECECEC", fg="#333333"
)
lbl_title.pack(pady=(30, 5))

lbl_subtitle = tk.Label(
    root, text="Скрипт автоматически найдет Excel-файл\nи удалит из него все сетевые компании.", 
    font=("Helvetica", 12), bg="#ECECEC", fg="#666666"
)
lbl_subtitle.pack(pady=(0, 25))

# Стартовая кнопка запуска
btn = tk.Button(
    root, text="Запустить очистку", command=clean_franchises, 
    font=("Helvetica", 14)
)
btn.pack(ipadx=10, ipady=3)

root.mainloop()
