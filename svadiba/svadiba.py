import asyncio
import urllib.parse
import openpyxl
import os
import re
from playwright.async_api import async_playwright, Error as PlaywrightError

def get_firm_id(url):
    if not url:
        return None
    match = re.search(r'/firm/(\d+)', str(url))
    return match.group(1) if match else None

async def bypass_museum(page):
    if "museum" in page.url:
        try:
            btn = page.get_by_text("Пропустить обновление")
            if await btn.count() > 0:
                await btn.first.click()
                await page.wait_for_timeout(1000)
        except Exception:
            pass

def generate_grid(lat_min, lat_max, lon_min, lon_max, steps):
    lat_step = (lat_max - lat_min) / steps
    lon_step = (lon_max - lon_min) / steps
    points = []
    
    grid_districts = [
        ["Внуково, Московский", "Теплый Стан, Коньково", "Бутово, Ясенево, Чертаново Юж.", "Бирюлево, Царицыно", "Зябликово, Братеево, Капотня"],
        ["Очаково, Солнцево", "Пр-т Вернадского, Раменки", "Чертаново Сев., Зюзино", "Нагатино, Печатники, Текстильщики", "Марьино, Люблино, Кузьминки"],
        ["Кунцево, Крылатское", "Филевский парк, Хамовники", "ЦАО (Арбат, Якиманка, Замоскворечье)", "Таганский, Басманный, Лефортово", "Перово, Новогиреево, Рязанский"],
        ["Строгино, Митино", "Хорошево-Мневники, Сокол", "Тверской, Пресня, Марьина Роща", "Сокольники, Алексеевский", "Измайлово, Гольяново"],
        ["Куркино, Сев. Тушино", "Ховрино, Головинский", "Отрадное, Коптево, Тимирязевский", "ВДНХ, Останкино, Свиблово", "Медведково, Бабушкинский"]
    ]

    for i in range(steps):
        for j in range(steps):
            lat = round(lat_min + lat_step * (i + 0.5), 6)
            lon = round(lon_min + lon_step * (j + 0.5), 6)
            
            col_letter = chr(65 + j)
            row_num = steps - i
            
            if steps == 5:
                districts = grid_districts[i][j]
                sector_name = f"Сектор {col_letter}{row_num} ({districts})"
            else:
                sector_name = f"Сектор {col_letter}{row_num}"
                
            points.append((lat, lon, sector_name))
            
    points.sort(key=lambda x: x[2])
    return points

async def process_firm(context, firm_id, url, ws, wb, file_path, lock, semaphore, state):
    async with semaphore:
        page = None
        try:
            page = await context.new_page()
            await page.route("**/*", lambda route: route.abort() 
                if route.request.resource_type in ["image", "stylesheet", "media", "font"] 
                else route.continue_())

            await page.goto(url, wait_until="domcontentloaded", timeout=25000)
            await bypass_museum(page)

            h1 = page.locator('h1')
            title = (await h1.first.inner_text()).split('\n')[0].strip() if await h1.count() > 0 else "Без названия"

            geo = page.locator('a[href*="/geo/"]')
            address = (await geo.first.inner_text()).replace('\n', ', ').strip() if await geo.count() > 0 else "Не указан"

            await page.evaluate("""() => {
                const btns = Array.from(document.querySelectorAll('button, div, span'));
                btns.forEach(b => {
                    const t = (b.innerText || '').toLowerCase();
                    if (t.includes('показать телефон') || t.includes('соцсети') || t.includes('показать контакт') || t.includes('ещё контакт')) {
                        try { b.click(); } catch(e) {}
                    }
                });
            }""")
            await page.wait_for_timeout(500)

            tels = page.locator('a[href^="tel:"]')
            phones = []
            for i in range(await tels.count()):
                href = await tels.nth(i).get_attribute('href')
                if href:
                    num = href.replace('tel:', '').strip()
                    if num and num not in phones:
                        phones.append(num)
            phone_str = ", ".join(phones) if phones else "Не указан"

            raw_data = await page.evaluate("""() => {
                let results = [];
                document.querySelectorAll('a').forEach(a => { if (a.href) results.push(a.href); });
                document.querySelectorAll('div, span, a').forEach(el => {
                    let txt = el.innerText;
                    if (txt) {
                        txt = txt.trim();
                        if (txt.length > 4 && txt.length < 70 && !txt.includes('\\n') && !txt.includes(' ')) {
                            results.push(txt);
                        }
                    }
                });
                return Array.from(new Set(results));
            }""")

            websites = []
            bad_domains = ["2gis", "yandex", "google", "flamp", "apple", "play.google", "otello", "restoclub", "tomesto", "zoon", "sbermarket", "megamarket", "delivery-club", "eda.yandex", "mos.ru", "gosuslugi", "w3.org", "github.com", "yoo.money"]
            good_indicators = [".ru", ".com", ".рф", ".org", ".net", ".info", ".su", ".pro", ".moscow", ".agency", ".media", ".digital", "vk.com", "t.me", "wa.me", "taplink.cc"]

            for val in raw_data:
                val_lower = val.lower()
                if "redirect?url=" in val_lower:
                    parsed = urllib.parse.parse_qs(urllib.parse.urlparse(val).query)
                    if "url" in parsed:
                        val = urllib.parse.unquote(parsed["url"][0])
                        val_lower = val.lower()

                if any(ind in val_lower for ind in good_indicators) and not any(bd in val_lower for bd in bad_domains):
                    if "@" not in val:
                        if not val_lower.startswith("http"):
                            val = "https://" + val
                        if val not in websites:
                            websites.append(val)

            real_sites = [w for w in websites if not any(s in w.lower() for s in ['vk.com', 't.me', 'instagram.com', 'wa.me', 'whatsapp.com'])]
            socials = [w for w in websites if any(s in w.lower() for s in ['vk.com', 't.me', 'instagram.com', 'wa.me', 'whatsapp.com'])]
            site_str = real_sites[0] if real_sites else (socials[0] if socials else "Нет сайта")

            async with lock:
                state["count"] += 1
                curr_no = state["count"]
                ws.append([curr_no, title, address, phone_str, site_str, url])
                wb.save(file_path)
                print(f"[{curr_no}] {title} | {address} | {phone_str} | {site_str}")

        except (asyncio.CancelledError, PlaywrightError):
            pass
        except Exception:
            pass
        finally:
            if page:
                try:
                    await page.close()
                except Exception:
                    pass

async def main():
    search_queries = [
        "свадебное агентство", 
        "организация свадеб", 
        "свадебный салон", 
        "свадебный организатор", 
        "свадебный декор",
        "оформление свадеб",
        "свадебный фотограф",
        "ведущий на свадьбу",
        "банкетный зал для свадьбы"
    ]
    city = "moscow"
    file_path = "свадебные_агентства.xlsx"
    progress_file = "progress_wedding.txt"
    sheet_title = "Свадебные агентства"
    
    LAT_MIN, LAT_MAX = 55.55, 55.92
    LON_MIN, LON_MAX = 37.35, 37.85
    GRID_STEPS = 5 
    ZOOM = 14  
    MAX_PAGES_PER_CELL = 10  # Увеличено для максимального охвата в центрах активности
    CONCURRENCY_LIMIT = 4

    saved_ids = set()
    lock = asyncio.Lock()
    semaphore = asyncio.Semaphore(CONCURRENCY_LIMIT)
    
    if os.path.exists(file_path):
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if len(row) >= 6 and row[5]:
                firm_id = get_firm_id(row[5])
                if firm_id:
                    saved_ids.add(firm_id)
        results_count = ws.max_row - 1
        print(f"Файл найден. Уже собрано уникальных карточек: {len(saved_ids)}")
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_title
        ws.append(["№", "Название", "Адрес", "Телефон", "Сайт", "URL"])
        results_count = 0

    done_sectors = set()
    if os.path.exists(progress_file):
        with open(progress_file, "r", encoding="utf-8") as f:
            done_sectors = set(line.strip() for line in f if line.strip())
        print(f"Загружен прогресс: пропущено уже готовых секторов: {len(done_sectors)}")

    state = {"count": results_count}
    grid_points = generate_grid(LAT_MIN, LAT_MAX, LON_MIN, LON_MAX, GRID_STEPS)

    async with async_playwright() as p:
        browser = await p.chromium.launch(
            headless=True, 
            args=["--disable-blink-features=AutomationControlled"]
        )
        context = await browser.new_context(
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/127.0.0.0 Safari/537.36",
            viewport={"width": 1280, "height": 720}
        )
        
        main_page = await context.new_page()

        try:
            for q_idx, search_query in enumerate(search_queries, 1):
                print(f"\n==================================================")
                print(f"🔎 Запрос [{q_idx}/{len(search_queries)}]: '{search_query}'")
                print(f"==================================================")
                encoded_query = urllib.parse.quote(search_query)

                for cell_idx, (lat, lon, sector_name) in enumerate(grid_points, 1):
                    progress_key = f"{search_query}|{sector_name}"

                    if progress_key in done_sectors:
                        print(f"⏩ Пропущен (уже готов): {search_query} | {sector_name}")
                        continue

                    print(f"\n📍 [{search_query}] [{cell_idx}/{len(grid_points)}]: {sector_name}")

                    for page_num in range(1, MAX_PAGES_PER_CELL + 1):
                        search_url = f"https://2gis.ru/{city}/search/{encoded_query}/page/{page_num}?m={lon}%2C{lat}%2F{ZOOM}"
                        
                        try:
                            await main_page.goto(search_url, wait_until="domcontentloaded", timeout=30000)
                            await bypass_museum(main_page)
                            await main_page.wait_for_timeout(1000)

                            # Прокрутка списка результатов для подгрузки динамического контента (Lazy Loading)
                            for _ in range(4):
                                await main_page.mouse.wheel(0, 3000)
                                await main_page.wait_for_timeout(500)

                        except Exception:
                            continue

                        cards = main_page.locator('a[href*="/firm/"]')
                        card_count = await cards.count()
                        if card_count == 0:
                            break

                        tasks = []
                        for i in range(card_count):
                            href = await cards.nth(i).get_attribute('href')
                            if href:
                                firm_id = get_firm_id(href)
                                if firm_id and firm_id not in saved_ids:
                                    clean = href.split('?')[0]
                                    full_url = clean if clean.startswith('http') else f"https://2gis.ru{clean}"
                                    
                                    saved_ids.add(firm_id) 
                                    tasks.append(process_firm(context, firm_id, full_url, ws, wb, file_path, lock, semaphore, state))

                        if tasks:
                            await asyncio.gather(*tasks, return_exceptions=True)

                    done_sectors.add(progress_key)
                    with open(progress_file, "a", encoding="utf-8") as f:
                        f.write(f"{progress_key}\n")

        except asyncio.CancelledError:
            pass
        finally:
            try:
                await browser.close()
            except Exception:
                pass

if __name__ == "__main__":
    try:
        asyncio.run(main())
    except KeyboardInterrupt:
        print("\n\n[!] Парсинг остановлен пользователем (Ctrl+C). Прогресс сохранен!")
